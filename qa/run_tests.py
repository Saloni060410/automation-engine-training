import sys
import os
import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from components.validator import validate_invoice
from components.pdf_parser import parse_invoice

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED   = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
HEADER_FILL = PatternFill("solid", fgColor="1F3864")

FUTURE = (datetime.date.today() + datetime.timedelta(days=30)).isoformat()
PAST   = (datetime.date.today() - datetime.timedelta(days=5)).isoformat()


def run(fn, *args, **kwargs):
    try:
        result = fn(*args, **kwargs)
        return str(result), None
    except Exception as e:
        return None, str(e)


TESTS = [
    # --- Happy path ---
    {
        "id": "TC001", "desc": "Valid invoice – Vendor A, INR, future date",
        "input": "vendor=Vendor A, amount=10000, currency=INR, due_date=future",
        "expected": "valid=True, errors=[]",
        "fn": lambda: validate_invoice({"vendor": "Vendor A", "amount": 10000, "currency": "INR", "due_date": FUTURE}),
        "pass_check": lambda r: r["valid"] is True,
    },
    {
        "id": "TC002", "desc": "Valid invoice – Vendor B, USD, future date",
        "input": "vendor=Vendor B, amount=25000, currency=USD, due_date=future",
        "expected": "valid=True, errors=[]",
        "fn": lambda: validate_invoice({"vendor": "Vendor B", "amount": 25000, "currency": "USD", "due_date": FUTURE}),
        "pass_check": lambda r: r["valid"] is True,
    },
    {
        "id": "TC003", "desc": "Valid invoice – Vendor C at exact credit limit (75000)",
        "input": "vendor=Vendor C, amount=75000, currency=INR, due_date=future",
        "expected": "valid=True, errors=[]",
        "fn": lambda: validate_invoice({"vendor": "Vendor C", "amount": 75000, "currency": "INR", "due_date": FUTURE}),
        "pass_check": lambda r: r["valid"] is True,
    },
    {
        "id": "TC004", "desc": "Valid invoice – Vendor D, small amount",
        "input": "vendor=Vendor D, amount=500, currency=USD, due_date=future",
        "expected": "valid=True, errors=[]",
        "fn": lambda: validate_invoice({"vendor": "Vendor D", "amount": 500, "currency": "USD", "due_date": FUTURE}),
        "pass_check": lambda r: r["valid"] is True,
    },
    {
        "id": "TC005", "desc": "Valid vendor – Vendor A is in APPROVED_VENDORS",
        "input": "vendor=Vendor A",
        "expected": "No 'Vendor not approved' error",
        "fn": lambda: validate_invoice({"vendor": "Vendor A", "amount": 100, "currency": "INR", "due_date": FUTURE}),
        "pass_check": lambda r: "Vendor not approved" not in r["errors"],
    },
    {
        "id": "TC006", "desc": "Valid currency – INR accepted",
        "input": "currency=INR",
        "expected": "No currency error",
        "fn": lambda: validate_invoice({"vendor": "Vendor A", "amount": 100, "currency": "INR", "due_date": FUTURE}),
        "pass_check": lambda r: "Currency must be INR or USD" not in r["errors"],
    },
    {
        "id": "TC007", "desc": "Valid currency – USD accepted",
        "input": "currency=USD",
        "expected": "No currency error",
        "fn": lambda: validate_invoice({"vendor": "Vendor A", "amount": 100, "currency": "USD", "due_date": FUTURE}),
        "pass_check": lambda r: "Currency must be INR or USD" not in r["errors"],
    },
    {
        "id": "TC008", "desc": "Future due date accepted",
        "input": f"due_date={FUTURE}",
        "expected": "No due date error",
        "fn": lambda: validate_invoice({"vendor": "Vendor A", "amount": 100, "currency": "INR", "due_date": FUTURE}),
        "pass_check": lambda r: "Due date is not in the future" not in r["errors"],
    },
    {
        "id": "TC009", "desc": "Amount below credit limit accepted",
        "input": "vendor=Vendor D, amount=1000 (limit=20000)",
        "expected": "No credit limit error",
        "fn": lambda: validate_invoice({"vendor": "Vendor D", "amount": 1000, "currency": "INR", "due_date": FUTURE}),
        "pass_check": lambda r: not any("credit limit" in e for e in r["errors"]),
    },
    {
        "id": "TC010", "desc": "Parse valid PDF – Invoice1.pdf",
        "input": "file=test_data/Invoice1.pdf",
        "expected": "Returns dict with vendor, amount fields",
        "fn": lambda: parse_invoice("test_data/Invoice1.pdf"),
        "pass_check": lambda r: isinstance(r, dict) and "vendor" in r,
    },
    # --- Edge cases ---
    {
        "id": "TC011", "desc": "Bad/corrupted PDF – non-existent file",
        "input": "file=test_data/fake_invoice.pdf",
        "expected": "Raises exception / returns error",
        "fn": lambda: parse_invoice("test_data/fake_invoice.pdf"),
        "pass_check": lambda r: False,  # always expects exception
        "expect_exception": True,
    },
    {
        "id": "TC012", "desc": "Missing vendor name – empty string",
        "input": "vendor=''",
        "expected": "valid=False, 'Vendor not approved' in errors",
        "fn": lambda: validate_invoice({"vendor": "", "amount": 100, "currency": "INR", "due_date": FUTURE}),
        "pass_check": lambda r: r["valid"] is False and "Vendor not approved" in r["errors"],
    },
    {
        "id": "TC013", "desc": "Invalid amount – negative (-500)",
        "input": "amount=-500",
        "expected": "valid=False, error about invalid amount",
        "fn": lambda: validate_invoice({"vendor": "Vendor A", "amount": -500, "currency": "INR", "due_date": FUTURE}),
        "pass_check": lambda r: r["valid"] is False,
    },
    {
        "id": "TC014", "desc": "Invalid amount – zero (0)",
        "input": "amount=0",
        "expected": "valid=False, error about invalid amount",
        "fn": lambda: validate_invoice({"vendor": "Vendor A", "amount": 0, "currency": "INR", "due_date": FUTURE}),
        "pass_check": lambda r: r["valid"] is False,
    },
    {
        "id": "TC015", "desc": "Currency not INR/USD – EUR rejected",
        "input": "currency=EUR",
        "expected": "valid=False, 'Currency must be INR or USD' in errors",
        "fn": lambda: validate_invoice({"vendor": "Vendor A", "amount": 100, "currency": "EUR", "due_date": FUTURE}),
        "pass_check": lambda r: r["valid"] is False and "Currency must be INR or USD" in r["errors"],
    },
]


def run_all():
    results = []
    for t in TESTS:
        expect_exc = t.get("expect_exception", False)
        try:
            result = t["fn"]()
            actual = str(result)
            if expect_exc:
                status = "Fail"
                actual = f"No exception raised. Got: {actual}"
            else:
                status = "Pass" if t["pass_check"](result) else "Fail"
        except Exception as e:
            actual = f"Exception: {e}"
            status = "Pass" if expect_exc else "Fail"

        results.append({**t, "actual": actual, "status": status})
        mark = "✓" if status == "Pass" else "✗"
        print(f"[{mark}] {t['id']} {t['desc'][:55]:<55} → {status}")

    return results


def save_xlsx(results):
    os.makedirs("qa", exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Matrix"

    headers = ["Test ID", "Description", "Input", "Expected Output", "Actual Output", "Status", "Severity"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    severity_map = {
        "TC011": "High", "TC012": "Medium",
        "TC013": "Critical", "TC014": "High", "TC015": "Medium",
    }

    for row, t in enumerate(results, 2):
        sev = severity_map.get(t["id"], "Low")
        values = [t["id"], t["desc"], t["input"], t["expected"], t["actual"], t["status"], sev]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True)

        status_cell = ws.cell(row=row, column=6)
        status_cell.fill = GREEN if t["status"] == "Pass" else RED

    col_widths = [10, 45, 35, 35, 50, 10, 10]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    path = "qa/test_matrix.xlsx"
    wb.save(path)
    print(f"\nSaved: {path}")
    return results


if __name__ == "__main__":
    print("Running 15 test cases...\n")
    results = run_all()
    save_xlsx(results)
    passed = sum(1 for r in results if r["status"] == "Pass")
    failed = sum(1 for r in results if r["status"] == "Fail")
    print(f"\nTotal: {len(results)} | Passed: {passed} | Failed: {failed}")
